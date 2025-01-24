
import json
import time
from collections import OrderedDict
import copy
import pandas
import os
import openpyxl
import pandas as pd
from openpyxl import load_workbook


# 将远端support中'sp': true的项在本地support中查找对比，为false的项不需要查找，因为本地的support文件中不存'sp': false的项
# 原有的 remote support key，雷达、配网WiFi频段、存储卡信息、供电模式、视频清晰度、压缩格式二维码、太阳能板、低功耗设备、夜视、卡录播放速度
# Android不包含video共9项，iOS包含video共10项，每项固定，是远端的优化点
remoteSupportKeyOfAndroid = ['radar', 'wifi', 'storage', 'power', 'compressedQrCode', 'equipSolarPanel', 'lowpowerdevice', 'nightVision', 'cardRecordSpeed']
remoteSupportKeyOfiOS = ['radar', 'wifi', 'storage', 'power', 'video', 'compressedQrCode', 'equipSolarPanel', 'lowpowerdevice', 'nightVision', 'cardRecordSpeed']

# 下沉后的 local support key，Android一共9项（不包含video），iOS一共10项（包含video），每项固定，为此次优化点
localSupportKeyOfAndroid = ['spRadar', 'spWifiBand', 'spStorage', 'spPower', 'spCompressedQrCode', 'spEquipSolarPanel', 'spLowPowerDevice', 'spNightVision', 'spCardRecordSpeed']
localSupportKeyOfiOS = ['spRadar', 'spWifi', 'spStorage', 'spPower', 'spVideo', 'spCompressedQrCode', 'spEquipSolarPanel', 'spLowPowerDevice', 'spNightVision', 'spCardRecordSpeed']


def openRemoteSupportFile(defaultPath='/Users/testmanzhang/Documents/data/'):
    """
    :param model: 远程dev_support.json文件中的每个设备的model，例如，"model": "AC9C3CA11"
    :param defaultPath:远程dev_support.json文件所放的位置
    :return:
    """
    with open(defaultPath + 'dev_support.json', 'r') as f:
        data = json.load(f)

    return data


def openLocalSupportFile(devModel, oSystem, defaultPath='/Users/testmanzhang/Documents/data/'):
    """
    :desc 开发同学提供的下沉之后的support文件，即放在本地的support文件。
    :param defaultPath: 文件的默认路径，即文件放的位置
    :param devModel: 设备型号，例如，c2e
    :param oSystem: Android 或者 iOS
    :return: 打开后的json文件
    """
    if oSystem == 'Android':
        filePath = defaultPath + 'devices/' + devModel.upper() + '/res/raw/'
        localSupportFileName = 'device_local_support_' + devModel.lower() + '.json'

    elif oSystem == 'iOS':
        filePath = defaultPath + 'iOS_Devices/' + devModel.upper() + '/'
        localSupportFileName = 'device_support.json'

    with open(filePath + localSupportFileName, 'r') as f:
        data = json.load(f)

    return data


def getDevicesModel():
    remoteSupports = openRemoteSupportFile()
    # 取出所有的设备型号，例如，"V8S"、"SV8S"..."C8E"
    devsModel = [remoteSupport["model"] for remoteSupport in remoteSupports]

    return devsModel


def getRemoteData(filePath, oSystem):
    global remoteSupportKeys

    # 此处不区分Android和iOS，远程文件是一个
    remoteSupports = openRemoteSupportFile()

    # 这里不读取已有的excel，因为，此次添加的数据跟读取的dataframe行数不一样，无法循环插入
    # df = pandas.read_excel(filePath, sheet_name='Sheet1')

    # 创建新的dataframe，用来一次性写入数据
    dfNewData = pandas.DataFrame({
        "Remote model": [],
        "Remote funcId": [],
        "Remote support": []
    })

    # 每行内容列表
    remoteModel = []
    remoteFuncId = []
    remoteSupport = []

    # 查找本次的改动项，Android一共9项（不包含video），iOS一共10项（包含video），参考wiki：https://wiki.glazero.com/pages/viewpage.action?pageId=26268592
    if oSystem == 'Android':
        remoteSupportKeys = remoteSupportKeyOfAndroid
    elif oSystem == 'iOS':
        remoteSupportKeys = remoteSupportKeyOfiOS

    for item in remoteSupports:
        # 每个型号对应一个item，获取对应的设备型号
        # df["Remote model"] = item["model"]
        # 按照support项遍历item["config"]
        for remoteSupportItem in remoteSupportKeys:
            supprtOfRemote_copy = copy.deepcopy(item)
            print(remoteSupportItem)

            # 在远程support文件中取出设备型号
            model = supprtOfRemote_copy["model"]
            for ii in supprtOfRemote_copy["config"]:
                if ii["funcId"] == remoteSupportItem:
                    funcId = ii["funcId"]
                    print(ii["funcId"])
                    # 不能用pop删除json项，返回值是删除的项
                    # aa = ii.pop("funcId")
                    del ii["funcId"]
                    print("删除funcId后的项：", ii)
                    # 删除项目后需要排序，不然remote和local对比不一致
                    iiAfterSorted = OrderedDict(sorted(ii.items()))
                    iiJson = json.dumps(iiAfterSorted)
                    # 此处保留，需要给剩余的json排序的话打开
                    # sortedData1 = OrderedDict(sorted(aa.items()))
                    # sortedData1 = aa
                    # print("sortedData1:", sortedData1)
                    # 写入excel
                    # df = pandas.DataFrame(
                    #    {"Model": [devModel], "Local funcId": [cc], "Local support": [ii]})
                    # 如果使用pandas.Series的话，会写入1行，后面的272行都空着，这样不太好
                    # df["Remote model"] = pandas.Series([xx])
                    # df["Remote funcId"] = pandas.Series([cc])
                    # df["Remote support"] = pandas.Series([ii])
                    # 每次都写入的话最后只写入一行，因为这是在已有excel中追加，新加入的1行跟原有的dataframe行数不一样，报错
                    # dfNewData["Remote model"] = [xx]
                    # dfNewData["Remote funcId"] = [cc]
                    # dfNewData["Remote support"] = [ii]
                    remoteModel.append(model)
                    remoteFuncId.append(funcId)
                    remoteSupport.append(iiJson)

    dfNewData["Remote model"] = remoteModel
    dfNewData["Remote funcId"] = remoteFuncId
    dfNewData["Remote support"] = remoteSupport

    with pandas.ExcelWriter(filePath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # 获取最大行数
        # max_row = writer.sheets['Sheet1'].max_row
        # startrow = max_row 此处不能从最大行处开始写，因为会空出前273行
        # 拼接的话是在原有excel的基础上即原有的行数上开始拼接，不能按预期写入
        # combiedData = pandas.concat([df, dfNew], ignore_index=True)
        dfNewData.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=0)


def getLocalData(filePath, oSystem):
    """
    :param filePath: 生成的excel文件
    :return: 将下沉到本地的每个设备型号（57个设备）的support项（Android 9项，iOS 10项），抽取出来放到excel中
    """
    global localSupportKeys

    # Android和iOS相同，都是从remote support中获取
    devsModel = getDevicesModel()

    # 读取已有的excel，按列插入数据
    df = pandas.read_excel(filePath, sheet_name='Sheet1')

    # 查找本次的改动项，Android一共9项（不包含video），iOS一共10项（包含video），参考wiki：https://wiki.glazero.com/pages/viewpage.action?pageId=26268592
    if oSystem == 'Android':
        localSupportKeys = localSupportKeyOfAndroid
    elif oSystem == 'iOS':
        localSupportKeys = localSupportKeyOfiOS

    for devModel in devsModel:
        print(devModel)

        # 打开对应型号的local support文件，此处区分Android和iOS
        supportOflocal = openLocalSupportFile(devModel, oSystem)

        for localSupportItem in localSupportKeys:
            supportOflocal_copy = copy.deepcopy(supportOflocal)

            print(localSupportItem)

            # 在下沉文件中取出设备型号，这样更准确一些，也可以用当前的devModel这个是从远程文件中获取到的
            model = supportOflocal_copy["model"]
            for ii in supportOflocal_copy["config"]:
                if ii["funcId"] == localSupportItem:
                    funcId = ii["funcId"]
                    print(ii["funcId"])
                    # aa = ii.pop("funcId")
                    del ii["funcId"]
                    print("删除funcId后的项：", ii)
                    # 删除项目后需要排序，不然remote和local对比不一致
                    iiAfterSorted = OrderedDict(sorted(ii.items()))
                    iiJson = json.dumps(iiAfterSorted)
                    # iiDict = dict(iiAfterSorted)
                    # 写入excel
                    # df = pandas.DataFrame(
                    #    {"Model": [devModel], "Local funcId": [cc], "Local support": [ii]})
                    df["Local model"] = [model]
                    df["Local funcId"] = [funcId]
                    df["Local support"] = [iiJson]
                    with pandas.ExcelWriter(filePath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        # 获取最大行数
                        max_row = writer.sheets['Sheet1'].max_row
                        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=max_row)


def compareOfRemoteAndLocal(filePath):
    df = pandas.read_excel(filePath, sheet_name='Sheet1')

    df["Compare result"] = df.apply(lambda row: "一致" if row["Remote support"] == row["Local support"] and row["Remote model"] == row["Local model"] else "不一致", axis=1)

    df.to_excel(filePath, sheet_name='Sheet1', index=False, header=True, engine='openpyxl')


def styleApply(content, colors, backGround=''):
    if content is not None and content in colors.keys():
        return "background-color: " + colors[content]
    return backGround


def styleColor(df, colors):
    return df.style.map(styleApply, colors=colors)


def highlightColumns(s):
    return ['background-color: lightblue' if i == 0 else '' for i in range(len(s))]


def creatResultExcel():
    # 创建结果excel文件
    filePath = "/Users/testmanzhang/Documents/data/Result_{}.xlsx".format(int(time.time()*1000))

    # 默认格式，创建表头
    df = pandas.DataFrame({"Remote model": [], "Remote funcId": [], "Remote support": [], "Local model": [], "Local funcId": [], "Local support": [], "Compare result": []})
    # df.to_excel(filePath)

    # writer = pandas.ExcelWriter(filePath, engine="openpyxl")
    # "#ff0000"红色、"#ffff00"黄色、"#1C1C1C"灰色、"#00EEEE"亮蓝色、"#1A1A1A"深灰色

    # colors = {"Model": "#ff0000", "Local funcId": "#ffff00", "Remote funcId": "#1C1C1C", "Local support": "#00EEEE", "Remote support": "#1A1A1A", "Compare result": "#ff0000"}

    # styleDf = styleColor(df, colors)

    # styleDf.to_excel(writer, index=False)

    # writer.close()

    # 设置表头背景
    # headerStyle = {"bgColor": "#00EEEE", "border": True}
    # df.style.set_table_styles([{
    #     "selector": "th",
    #     "props": [("background-color", headerStyle["bgColor"]), ("border", "1px solid black")]
    # }]).to_excel(filePath, index=True)

    styler = df.style.apply(highlightColumns, axis=1).set_properties(**{"border": "1px solid black", "text-align": "center"}).set_table_styles([{"selector": "thead", "props": [("font-weight", "bold")]}]).set_caption("My table")
    styler.to_excel(filePath, engine='openpyxl', index=False)

    return filePath


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    filePath = creatResultExcel()
    print(filePath)
    getLocalData(filePath, oSystem="Android")
    print("get local data done!\n")
    getRemoteData(filePath, oSystem="Android")
    print("get remote data done!")
    compareOfRemoteAndLocal(filePath)
    print("对比完成！")
